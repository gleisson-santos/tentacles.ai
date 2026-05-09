import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import calendar
import os

def generate_timesheet(year, month, output_path):
    # Create directory if not exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ponto {month}-{year}"

    # Colors and Styles
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid") # Dark Clilink
    header_font = Font(color="FFFFFF", bold=True)
    accent_fill = PatternFill(start_color="0077B5", end_color="0077B5", fill_type="solid") # Blue Clilink
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Header Row
    headers = ["Data", "Entrada", "Pausa (Início)", "Pausa (Fim)", "Saída", "Assinatura"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Days of the month
    num_days = calendar.monthrange(year, month)[1]
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        row = day + 1
        
        # Write date
        date_cell = ws.cell(row=row, column=1, value=current_date.strftime("%d/%m/%Y (%a)"))
        date_cell.border = thin_border
        
        # Style row (weekend highlight)
        if current_date.weekday() >= 5: # Saturday or Sunday
            fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        
        # Empty cells with borders
        for col in range(2, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_alignment

    # Adjust columns width
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions['F'].width = 30

    # Save
    wb.save(output_path)
    print(f"Planilha gerada com sucesso em: {output_path}")

if __name__ == "__main__":
    # Using May 2026 as per session context
    generate_timesheet(2026, 5, "outputs/planilhas/folha_de_ponto.xlsx")
